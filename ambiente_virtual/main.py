import requests
from openpyxl import Workbook
from fpdf import FPDF
import pyautogui

resposta = requests.get("https://economia.awesomeapi.com.br/json/last/USD-BRL")
dado = resposta.json()
valor_dolar = dado["USDBRL"]["bid"]

wb = Workbook()
ws = wb.active
ws["A1"] = "Cotação do Dólar"
ws["B1"] = float(valor_dolar)
wb.save("cotacao.xlsx")

pdf = FPDF()
pdf.add_page()
pdf.set_font("Helvetica", size=16)
pdf.cell(0, 10, f"Cotação atual do dólar: R$ {valor_dolar}", ln=True)
pdf.output("cotacao.pdf")

pyautogui.screenshot("tela.png")

print("Relatórios gerados com sucesso!")
