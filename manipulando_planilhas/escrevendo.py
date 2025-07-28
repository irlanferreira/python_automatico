from openpyxl import Workbook

arquivo = Workbook()

planilha_atual = arquivo.active
planilha_atual.title = "Produtos"

# planilha_atual['A1'] = 'Nome'
# planilha_atual['B1'] = "Preço"

# planilha_atual['A2'] = 'Maçã'
# planilha_atual['B2'] = 1.78

planilha_atual.append(['Nome', 'Preço'])
planilha_atual.append(['Maçã', 2.45])
planilha_atual.append(['Pão', 1.23])
planilha_atual['B3'] = 1.67

arquivo.create_sheet('Vendas')
planilha_vendas = arquivo['Vendas']

planilha_vendas.append(['Valor', 'Data'])
planilha_vendas.append([43.50, '29/07'])

print(arquivo.sheetnames)

arquivo.save('planilhas.xlsx')