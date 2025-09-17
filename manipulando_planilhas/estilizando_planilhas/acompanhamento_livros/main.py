from openpyxl import Workbook
from datetime import datetime
from openpyxl.styles import Alignment, PatternFill, Font, Border, Side

arquivo = Workbook()
planilha = arquivo.active
planilha.title = "Livros"

planilha['A1'] = "Diário de Leituras - Setembro 2025"
planilha.merge_cells('A1:D1')
planilha['A1'].alignment = Alignment(horizontal="center")
planilha['A1'].fill = PatternFill(fgColor="1F497D", fill_type="solid")
planilha['A1'].font = Font(bold=True, size=14, color="FFFFFF")

planilha.append(['Livro', 'Autor', 'Data de Início', 'Progresso'])
for celula in planilha[2]:
    celula.font = Font(color="FFFFFF", bold=True)
    celula.fill = PatternFill(fgColor="4f4f4f", fill_type="solid")

planilha.append(['Desmarketise-ze', 'João Branco', datetime(2025, 10, 4), 1])
planilha.append(['A vida Secreta dos Clientes', 'David S. Duncan', datetime(2025, 9, 20), 0.82])
planilha.append(['A vida Secreta dos Clientes', 'David S. Duncan', datetime(2025, 9, 20), 0.82])
planilha.append(['A vida Secreta dos Clientes', 'David S. Duncan', datetime(2025, 9, 20), 0.82])
planilha.append(['A vida Secreta dos Clientes', 'David S. Duncan', datetime(2025, 9, 20), 0.82])
planilha.append(['A vida Secreta dos Clientes', 'David S. Duncan', datetime(2025, 9, 20), 0.82])
planilha.append(['A vida Secreta dos Clientes', 'David S. Duncan', datetime(2025, 9, 20), 0.82])
planilha.append(['A vida Secreta dos Clientes', 'David S. Duncan', datetime(2025, 9, 20), 0.82])
planilha.append(['A vida Secreta dos Clientes', 'David S. Duncan', datetime(2025, 9, 20), 0.82])


for celula in planilha['C']:
    celula.number_format = "DD/MM/YY"
for celula in planilha['D']:
    celula.number_format = "0%"

for linha in planilha.iter_rows():
    for celula in linha:
        celula.border = Border(Side("thin"), Side("thin"), Side("thin"), Side("thin"))
        if celula.row <= 2:
            continue
        if celula.row % 2 == 0:
            celula.fill = PatternFill(fgColor="C9C9C9", fill_type="solid")

arquivo.save("livros.xlsx")