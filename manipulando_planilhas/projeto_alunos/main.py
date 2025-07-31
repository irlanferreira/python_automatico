from openpyxl import load_workbook, Workbook

arquivo_alunos = load_workbook('alunos.xlsx')
planilha_alunos = arquivo_alunos['Alunos']

alunos_aprovados = []
alunos_reprovados = []

soma_nota = 0
maior_nota = 0

for aluno in planilha_alunos.iter_rows(min_row=2, values_only=True):
    nome, curso, idade, nota_final, data_matricula = aluno
    soma_nota+=nota_final
    if nota_final > maior_nota:
        maior_nota = nota_final
        aluno_maior_nota = nome
    aluno_dicionario = {
                'nome': nome,
                'curso': curso,
                'idade': idade,
                'nota_final': nota_final,
                'data_matricula': data_matricula
            }
    if nota_final >= 7:
        alunos_aprovados.append(aluno_dicionario)
    else:
        alunos_reprovados.append(aluno_dicionario)

media_nota = soma_nota / (len(alunos_aprovados)+len(alunos_reprovados))

arquivo_aprovados = Workbook()
planilha_aprovados = arquivo_aprovados.active
planilha_aprovados.title = 'Alunos aprovados'

arquivo_reprovados = Workbook()
planilha_reprovados = arquivo_reprovados.active
planilha_reprovados.title = 'Alunos reprovados'

planilha_aprovados.append(['Nome', 'Curso', 'Idade', 'Nota Final', 'Data de Matrícula'])
for aluno in alunos_aprovados:
    planilha_aprovados.append(list(aluno.values()))

planilha_reprovados.append(['Nome', 'Curso', 'Idade', 'Nota Final', 'Data de Matrícula'])
for aluno in alunos_reprovados:
    planilha_reprovados.append(list(aluno.values()))


arquivo_aprovados.save('aprovados.xlsx')
arquivo_reprovados.save('reprovados.xlsx')

print(f"Número de aprovados: {len(alunos_aprovados)}")
print(f"Número de reprovados: {len(alunos_reprovados)}")
print(f"Média de notas: {media_nota:.2f}")
print(f"Maior nota: {aluno_maior_nota} - {maior_nota}")