from openpyxl import load_workbook

arquivo = load_workbook('alunos.xlsx')
planilha_alunos = arquivo['Alunos']

# print(planilha_alunos['B2'].value)
# print(planilha_alunos['D5'].value)
# print(planilha_alunos['E10'].value)

# notas = planilha_alunos['D']

# for celula in notas:
#     if celula.row == 1:
#         continue
#     if celula.value > 8:
#         print(celula.value)

for linha in planilha_alunos.iter_rows(min_row=2, values_only=True):
    nome, curso, idade, nota_final, data_matricula = linha
    print('='*40)
    print(f"""NOME: {nome}
CURSO: {curso}
IDADE: {idade}
NOTA FINAL: {nota_final}
DATA DE MATRÍCULA: {data_matricula.strftime("%d/%m/%Y")}""")