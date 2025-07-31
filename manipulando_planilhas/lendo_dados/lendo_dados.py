from openpyxl import load_workbook

arquivo = load_workbook('planilha_funcionarios.xlsx')
planilha_funcionarios = arquivo['Funcionários']

# salario_anthony = planilha_funcionarios['D6'].value
# print(salario_anthony)

# linha_13 = planilha_funcionarios[13]
# for celula in linha_13:
#     print(celula.value)

# linhas = planilha_funcionarios[7:12]
# for linha in linhas:
#     print("="*30)
#     for celula in linha:
#         print(celula.value)

# coluna_salario = planilha_funcionarios['D']
# print(coluna_salario)

for linha in planilha_funcionarios.iter_rows(values_only=True, min_row=2, max_row=20):
    print('-'*50)
    nome, departamento, idade, salario, data_admissao = linha
    print(f"""NOME: {nome}
DEPARTAMENTO: {departamento}
IDADE: {idade}
SALÁRIO: {salario}
DATA DE ADMISSÃO: {data_admissao.strftime("%d/%m/%Y")}""")